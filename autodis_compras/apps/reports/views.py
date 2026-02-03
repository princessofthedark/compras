"""
Views para el modulo de reportes.
Genera reportes dinamicos desde los datos de solicitudes y presupuestos.
Soporta exportacion a Excel (.xlsx) y PDF.
"""

import io
from decimal import Decimal
from django.db.models import Sum, Count
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import permissions, status
from rest_framework.views import APIView
from rest_framework.response import Response

from autodis_compras.apps.requests.models import PurchaseRequest
from autodis_compras.apps.budgets.models import Budget


APPROVED_STATUSES = [
    PurchaseRequest.APROBADA,
    PurchaseRequest.EN_PROCESO,
    PurchaseRequest.COMPRADA,
    PurchaseRequest.COMPLETADA,
]


def _build_expenses_queryset(year, month=None, area_id=None, cost_center_id=None, category_id=None):
    """Construye queryset filtrado de solicitudes aprobadas."""
    qs = PurchaseRequest.objects.filter(
        status__in=APPROVED_STATUSES,
        created_at__year=year,
    )
    if month:
        qs = qs.filter(created_at__month=month)
    if area_id:
        qs = qs.filter(cost_center__area_id=area_id)
    if cost_center_id:
        qs = qs.filter(cost_center_id=cost_center_id)
    if category_id:
        qs = qs.filter(category_id=category_id)
    return qs


class ExpensesByPeriodView(APIView):
    """Reporte de gastos por periodo (mes/trimestre/anio)."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        area_id = request.query_params.get('area')
        cost_center_id = request.query_params.get('cost_center')
        category_id = request.query_params.get('category')
        export = request.query_params.get('export')

        if not year:
            return Response({'error': 'El parametro year es obligatorio.'}, status=status.HTTP_400_BAD_REQUEST)

        qs = _build_expenses_queryset(year, month, area_id, cost_center_id, category_id)

        by_category = list(qs.values('category__name').annotate(
            total=Sum('estimated_amount'), count=Count('id'),
        ).order_by('-total'))

        by_cost_center = list(qs.values(
            'cost_center__code', 'cost_center__name'
        ).annotate(
            total=Sum('estimated_amount'), count=Count('id'),
        ).order_by('-total'))

        totals = qs.aggregate(
            total_estimated=Sum('estimated_amount'),
            total_actual=Sum('actual_amount'),
            total_count=Count('id'),
        )

        if export == 'excel':
            return self._export_excel(by_category, by_cost_center, totals, year, month)
        if export == 'pdf':
            return self._export_pdf(by_category, by_cost_center, totals, year, month)

        return Response({
            'filters': {'year': year, 'month': month, 'area': area_id, 'cost_center': cost_center_id, 'category': category_id},
            'totals': totals,
            'by_category': by_category,
            'by_cost_center': by_cost_center,
        })

    def _export_excel(self, by_category, by_cost_center, totals, year, month):
        import openpyxl
        from openpyxl.styles import Font, Alignment

        wb = openpyxl.Workbook()

        # Hoja: Por Categoria
        ws = wb.active
        ws.title = 'Por Categoria'
        ws.append(['Categoria', 'Total', 'Cantidad'])
        ws['A1'].font = ws['B1'].font = ws['C1'].font = Font(bold=True)
        for row in by_category:
            ws.append([row['category__name'], float(row['total'] or 0), row['count']])

        # Hoja: Por Centro de Costos
        ws2 = wb.create_sheet('Por Centro de Costos')
        ws2.append(['Codigo', 'Centro de Costos', 'Total', 'Cantidad'])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        for row in by_cost_center:
            ws2.append([row['cost_center__code'], row['cost_center__name'], float(row['total'] or 0), row['count']])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        period = f'{year}-{month:0>2}' if month else year
        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="gastos_periodo_{period}.xlsx"'
        return response

    def _export_pdf(self, by_category, by_cost_center, totals, year, month):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        period = f'{year}/{month:0>2}' if month else year
        elements.append(Paragraph(f'Reporte de Gastos - {period}', styles['Title']))
        elements.append(Spacer(1, 0.3 * inch))

        total_est = totals.get('total_estimated') or 0
        elements.append(Paragraph(f'Total estimado: ${float(total_est):,.2f} MXN | Solicitudes: {totals.get("total_count", 0)}', styles['Normal']))
        elements.append(Spacer(1, 0.3 * inch))

        # Tabla por categoria
        elements.append(Paragraph('Gastos por Categoria', styles['Heading2']))
        data = [['Categoria', 'Total', 'Cantidad']]
        for row in by_category:
            data.append([row['category__name'], f'${float(row["total"] or 0):,.2f}', str(row['count'])])

        if len(data) > 1:
            t = Table(data, colWidths=[3 * inch, 2 * inch, 1.5 * inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ]))
            elements.append(t)

        doc.build(elements)
        buf.seek(0)

        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="gastos_periodo_{period}.pdf"'
        return response


class BudgetComparisonView(APIView):
    """Reporte de comparacion presupuesto vs gasto real."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        export = request.query_params.get('export')

        if not year:
            return Response({'error': 'El parametro year es obligatorio.'}, status=status.HTTP_400_BAD_REQUEST)

        budgets_qs = Budget.objects.select_related('cost_center', 'category').filter(year=year)
        if month:
            budgets_qs = budgets_qs.filter(month=month)

        results = []
        for budget in budgets_qs:
            spent = budget.get_spent_amount()
            results.append({
                'cost_center': budget.cost_center.code,
                'cost_center_name': budget.cost_center.name,
                'category': budget.category.name,
                'year': budget.year,
                'month': budget.month,
                'budgeted': budget.amount,
                'spent': spent,
                'available': budget.amount - spent,
                'utilization_pct': float(budget.get_utilization_percentage()),
                'exceeded': spent > budget.amount,
            })

        if export == 'excel':
            return self._export_excel(results, year, month)
        if export == 'pdf':
            return self._export_pdf(results, year, month)

        return Response({'filters': {'year': year, 'month': month}, 'results': results})

    def _export_excel(self, results, year, month):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Comparativo'
        headers = ['Centro Costos', 'Categoria', 'Mes', 'Presupuestado', 'Gastado', 'Disponible', '% Utilizacion', 'Excedido']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        for i, row in enumerate(results, start=2):
            ws.append([
                row['cost_center'], row['category'], row['month'],
                float(row['budgeted']), float(row['spent']), float(row['available']),
                row['utilization_pct'], 'SI' if row['exceeded'] else 'NO',
            ])
            if row['exceeded']:
                for cell in ws[i]:
                    cell.fill = red_fill

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        period = f'{year}-{month:0>2}' if month else year
        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="comparativo_presupuesto_{period}.xlsx"'
        return response

    def _export_pdf(self, results, year, month):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []

        period = f'{year}/{month:0>2}' if month else year
        elements.append(Paragraph(f'Comparativo de Presupuesto - {period}', styles['Title']))
        elements.append(Spacer(1, 0.3 * inch))

        data = [['Centro', 'Categoria', 'Mes', 'Presupuestado', 'Gastado', 'Disponible', '% Uso']]
        for row in results:
            data.append([
                row['cost_center'], row['category'], str(row['month']),
                f'${float(row["budgeted"]):,.2f}', f'${float(row["spent"]):,.2f}',
                f'${float(row["available"]):,.2f}', f'{row["utilization_pct"]:.1f}%',
            ])

        if len(data) > 1:
            t = Table(data)
            style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ]
            for i, row in enumerate(results, start=1):
                if row['exceeded']:
                    style.append(('BACKGROUND', (0, i), (-1, i), colors.Color(1, 0.8, 0.8)))
            t.setStyle(TableStyle(style))
            elements.append(t)

        doc.build(elements)
        buf.seek(0)

        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="comparativo_presupuesto_{period}.pdf"'
        return response


class ExpensesByEmployeeView(APIView):
    """Reporte de gastos por empleado."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        export = request.query_params.get('export')

        if not year:
            return Response({'error': 'El parametro year es obligatorio.'}, status=status.HTTP_400_BAD_REQUEST)

        qs = PurchaseRequest.objects.filter(
            status__in=APPROVED_STATUSES, created_at__year=year,
        )
        if month:
            qs = qs.filter(created_at__month=month)

        by_employee = list(qs.values(
            'requester__first_name', 'requester__last_name', 'requester__email',
            'requester__area__name',
        ).annotate(
            total=Sum('estimated_amount'), count=Count('id'),
        ).order_by('-total'))

        if export == 'excel':
            return self._export_excel(by_employee, year, month)

        return Response({'filters': {'year': year, 'month': month}, 'results': by_employee})

    def _export_excel(self, results, year, month):
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Gastos por Empleado'
        ws.append(['Nombre', 'Apellido', 'Email', 'Area', 'Total', 'Solicitudes'])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in results:
            ws.append([
                row['requester__first_name'], row['requester__last_name'],
                row['requester__email'], row['requester__area__name'],
                float(row['total'] or 0), row['count'],
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        period = f'{year}-{month:0>2}' if month else year
        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="gastos_empleado_{period}.xlsx"'
        return response


class TopSuppliersView(APIView):
    """Reporte de proveedores mas utilizados."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        year = request.query_params.get('year')
        export = request.query_params.get('export')

        qs = PurchaseRequest.objects.filter(
            status__in=[PurchaseRequest.COMPRADA, PurchaseRequest.COMPLETADA],
        ).exclude(actual_supplier='')

        if year:
            qs = qs.filter(created_at__year=year)

        by_supplier = list(qs.values('actual_supplier').annotate(
            total=Sum('actual_amount'), count=Count('id'),
        ).order_by('-total')[:20])

        if export == 'excel':
            return self._export_excel(by_supplier, year)

        return Response({'filters': {'year': year}, 'results': by_supplier})

    def _export_excel(self, results, year):
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Top Proveedores'
        ws.append(['Proveedor', 'Total', 'Compras'])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in results:
            ws.append([row['actual_supplier'], float(row['total'] or 0), row['count']])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="proveedores_{year or "todos"}.xlsx"'
        return response


class DashboardSummaryView(APIView):
    """Resumen general para el dashboard."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        now = timezone.now()

        base_qs = PurchaseRequest.objects.all()
        if not (user.is_finance() or user.is_general_director()):
            if user.is_manager():
                base_qs = base_qs.filter(requester__area=user.area)
            else:
                base_qs = base_qs.filter(requester=user)

        pending_manager = base_qs.filter(status=PurchaseRequest.PENDIENTE_GERENTE).count()
        pending_finance = base_qs.filter(status=PurchaseRequest.APROBADA_POR_GERENTE).count()
        in_process = base_qs.filter(status=PurchaseRequest.EN_PROCESO).count()
        completed_this_month = base_qs.filter(
            status=PurchaseRequest.COMPLETADA,
            created_at__year=now.year, created_at__month=now.month,
        ).count()

        monthly_spend = base_qs.filter(
            status__in=APPROVED_STATUSES,
            created_at__year=now.year, created_at__month=now.month,
        ).aggregate(total=Sum('estimated_amount'))['total'] or Decimal('0.00')

        return Response({
            'pending_manager_approval': pending_manager,
            'pending_finance_approval': pending_finance,
            'in_process': in_process,
            'completed_this_month': completed_this_month,
            'monthly_spend': monthly_spend,
        })
